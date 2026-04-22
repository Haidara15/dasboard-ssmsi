"""
Génère un fichier exemple_import.xlsx dans le dossier racine du projet.
Contient un échantillon de chaque type de graphique sur une seule feuille.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from pathlib import Path

HEADERS = [
    # --- COLONNES OBLIGATOIRES (template minimal stable) ---
    "graphique_code",
    "thematique_code",
    "type_graphique",
    "serie_code",
    "code_x",
    "valeur",
]

# (graphique_code, thematique_code, type_graphique, serie_code, code_x, valeur, 
#  graphique_titre, thematique_titre, serie_libelle, libelle_x, titre_axe_x, titre_axe_y, source)

ROWS = [
    # =========================================================
    # 1. COLONNES simples
    # =========================================================
    ("colonnes-simple", "violences", "colonnes", "nombre", "2019", 809,
     "Homicides enregistrés", "Violences", "Nombre", "2019", "Année", "Nombre", "SSMSI"),
    ("colonnes-simple", "violences", "colonnes", "nombre", "2020", 863,
     "Homicides enregistrés", "Violences", "Nombre", "2020", "Année", "Nombre", "SSMSI"),
    ("colonnes-simple", "violences", "colonnes", "nombre", "2021", 887,
     "Homicides enregistrés", "Violences", "Nombre", "2021", "Année", "Nombre", "SSMSI"),
    ("colonnes-simple", "violences", "colonnes", "nombre", "2022", 910,
     "Homicides enregistrés", "Violences", "Nombre", "2022", "Année", "Nombre", "SSMSI"),
    ("colonnes-simple", "violences", "colonnes", "nombre", "2023", 884,
     "Homicides enregistrés", "Violences", "Nombre", "2023", "Année", "Nombre", "SSMSI"),

    # =========================================================
    # 2. COLONNES groupées (plusieurs séries)
    # =========================================================
    ("colonnes-groupees", "violences", "colonnes", "hommes", "2021", 620,
     "Victimes par sexe", "Violences", "Hommes", "2021", "Année", "Nombre", "SSMSI"),
    ("colonnes-groupees", "violences", "colonnes", "hommes", "2022", 641,
     "Victimes par sexe", "Violences", "Hommes", "2022", "Année", "Nombre", "SSMSI"),
    ("colonnes-groupees", "violences", "colonnes", "hommes", "2023", 612,
     "Victimes par sexe", "Violences", "Hommes", "2023", "Année", "Nombre", "SSMSI"),
    ("colonnes-groupees", "violences", "colonnes", "femmes", "2021", 267,
     "Victimes par sexe", "Violences", "Femmes", "2021", "Année", "Nombre", "SSMSI"),
    ("colonnes-groupees", "violences", "colonnes", "femmes", "2022", 269,
     "Victimes par sexe", "Violences", "Femmes", "2022", "Année", "Nombre", "SSMSI"),
    ("colonnes-groupees", "violences", "colonnes", "femmes", "2023", 272,
     "Victimes par sexe", "Violences", "Femmes", "2023", "Année", "Nombre", "SSMSI"),

    # =========================================================
    # 3. COLONNES EMPILEES
    # =========================================================
    ("colonnes-empilees", "violences", "colonnes_empilees", "moins30", "2021", 210,
     "Victimes par tranche d'âge", "Violences", "Moins de 30 ans", "2021", "Année", "Nombre", "SSMSI"),
    ("colonnes-empilees", "violences", "colonnes_empilees", "moins30", "2022", 225,
     "Victimes par tranche d'âge", "Violences", "Moins de 30 ans", "2022", "Année", "Nombre", "SSMSI"),
    ("colonnes-empilees", "violences", "colonnes_empilees", "30a60", "2021", 480,
     "Victimes par tranche d'âge", "Violences", "30 à 60 ans", "2021", "Année", "Nombre", "SSMSI"),
    ("colonnes-empilees", "violences", "colonnes_empilees", "30a60", "2022", 495,
     "Victimes par tranche d'âge", "Violences", "30 à 60 ans", "2022", "Année", "Nombre", "SSMSI"),
    ("colonnes-empilees", "violences", "colonnes_empilees", "plus60", "2021", 197,
     "Victimes par tranche d'âge", "Violences", "Plus de 60 ans", "2021", "Année", "Nombre", "SSMSI"),
    ("colonnes-empilees", "violences", "colonnes_empilees", "plus60", "2022", 190,
     "Victimes par tranche d'âge", "Violences", "Plus de 60 ans", "2022", "Année", "Nombre", "SSMSI"),

    # =========================================================
    # 4. LIGNE
    # =========================================================
    ("ligne-simple", "violences", "ligne", "taux", "2019", 1.21,
     "Taux d'homicides pour 100 000 hab.", "Violences", "Taux", "2019", "Année", "Taux", "SSMSI"),
    ("ligne-simple", "violences", "ligne", "taux", "2020", 1.29,
     "Taux d'homicides pour 100 000 hab.", "Violences", "Taux", "2020", "Année", "Taux", "SSMSI"),
    ("ligne-simple", "violences", "ligne", "taux", "2021", 1.32,
     "Taux d'homicides pour 100 000 hab.", "Violences", "Taux", "2021", "Année", "Taux", "SSMSI"),
    ("ligne-simple", "violences", "ligne", "taux", "2022", 1.35,
     "Taux d'homicides pour 100 000 hab.", "Violences", "Taux", "2022", "Année", "Taux", "SSMSI"),

    # =========================================================
    # 5. MIXTE colonnes + ligne (axe secondaire)
    # =========================================================
    ("mixte-axe-double", "violences", "colonnes", "nombre", "2019", 809,
     "Nombre et évolution (%)", "Violences", "Nombre", "2019", "Année", "Nombre", "SSMSI"),
    ("mixte-axe-double", "violences", "colonnes", "nombre", "2020", 863,
     "Nombre et évolution (%)", "Violences", "Nombre", "2020", "Année", "Nombre", "SSMSI"),
    ("mixte-axe-double", "violences", "colonnes", "nombre", "2021", 887,
     "Nombre et évolution (%)", "Violences", "Nombre", "2021", "Année", "Nombre", "SSMSI"),
    ("mixte-axe-double", "violences", "colonnes", "evolution", "2019", 0,
     "Nombre et évolution (%)", "Violences", "Évolution (%)", "2019", "Année", "Évolution (%)", "SSMSI"),
    ("mixte-axe-double", "violences", "colonnes", "evolution", "2020", 6.7,
     "Nombre et évolution (%)", "Violences", "Évolution (%)", "2020", "Année", "Évolution (%)", "SSMSI"),
    ("mixte-axe-double", "violences", "colonnes", "evolution", "2021", 2.8,
     "Nombre et évolution (%)", "Violences", "Évolution (%)", "2021", "Année", "Évolution (%)", "SSMSI"),

    # =========================================================
    # 6. AIRE
    # =========================================================
    ("aire-simple", "violences", "aire", "cumul", "2019", 809,
     "Cumul des victimes", "Violences", "Cumul", "2019", "Année", "Nombre", "SSMSI"),
    ("aire-simple", "violences", "aire", "cumul", "2020", 1672,
     "Cumul des victimes", "Violences", "Cumul", "2020", "Année", "Nombre", "SSMSI"),

    # =========================================================
    # 7. BARRES
    # =========================================================
    ("barres-simple", "violences", "barres", "nombre", "idf", 258,
     "Homicides par région", "Violences", "Nombre", "Île-de-France", "Région", "Nombre", "SSMSI"),
    ("barres-simple", "violences", "barres", "nombre", "paca", 134,
     "Homicides par région", "Violences", "Nombre", "PACA", "Région", "Nombre", "SSMSI"),
    ("barres-simple", "violences", "barres", "nombre", "auv-ra", 134,
     "Homicides par région", "Violences", "Nombre", "Auvergne-Rhône-Alpes", "Région", "Nombre", "SSMSI"),

    # =========================================================
    # 8. BARRES EMPILEES
    # =========================================================
    ("barres-empilees", "violences", "barres_empilees", "hommes", "idf", 220,
     "Homicides par région et sexe", "Violences", "Hommes", "Île-de-France", "Région", "Nombre", "SSMSI"),
    ("barres-empilees", "violences", "barres_empilees", "hommes", "paca", 130,
     "Homicides par région et sexe", "Violences", "Hommes", "PACA", "Région", "Nombre", "SSMSI"),
    ("barres-empilees", "violences", "barres_empilees", "femmes", "idf", 92,
     "Homicides par région et sexe", "Violences", "Femmes", "Île-de-France", "Région", "Nombre", "SSMSI"),
    ("barres-empilees", "violences", "barres_empilees", "femmes", "paca", 57,
     "Homicides par région et sexe", "Violences", "Femmes", "PACA", "Région", "Nombre", "SSMSI"),

    # =========================================================
    # 9. CAMEMBERT
    # =========================================================
    ("camembert-repartition", "violences", "camembert", "arme-feu", "arme-feu", 312,
     "Répartition par type d'acte", "Violences", "Arme à feu", "Arme à feu", "", "", "SSMSI"),
    ("camembert-repartition", "violences", "camembert", "arme-blanche", "arme-blanche", 298,
     "Répartition par type d'acte", "Violences", "Arme blanche", "Arme blanche", "", "", "SSMSI"),
    ("camembert-repartition", "violences", "camembert", "autres", "autres", 274,
     "Répartition par type d'acte", "Violences", "Autres", "Autres", "", "", "SSMSI"),

    # =========================================================
    # 10. ANNEAU
    # =========================================================
    ("anneau-repartition", "violences", "anneau", "hommes", "hommes", 641,
     "Répartition par sexe de la victime", "Violences", "Hommes", "Hommes", "", "", "SSMSI"),
    ("anneau-repartition", "violences", "anneau", "femmes", "femmes", 243,
     "Répartition par sexe de la victime", "Violences", "Femmes", "Femmes", "", "", "SSMSI"),

    # =========================================================
    # 11. VALEUR CLE
    # =========================================================
    ("valeur-cle-total", "violences", "valeur_cle", "total", "2023", 884,
     "Total 2023", "Violences", "Total", "2023", "", "", "SSMSI"),
]


def generer():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "import"

    # --- En-têtes ---
    header_fill = PatternFill("solid", fgColor="1a56db")
    header_font = Font(bold=True, color="FFFFFF")

    for col_idx, h in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # --- Couleurs alternées par graphique ---
    couleurs = ["F0F4FF", "FFF4F0", "F0FFF4", "FFF9F0", "F8F0FF", "F0FFFF", "FFFFF0", "FFF0F8", "F5F5F5", "E8F8E8", "FFF0E0"]
    couleur_par_graphique = {}

    for row_idx, row in enumerate(ROWS, start=2):
        graphique_code = row[0]
        if graphique_code and graphique_code not in couleur_par_graphique:
            couleur_par_graphique[graphique_code] = couleurs[len(couleur_par_graphique) % len(couleurs)]

        # Déterminer la couleur de fond
        code_ref = graphique_code if graphique_code else list(couleur_par_graphique.keys())[-1]
        fill_color = couleur_par_graphique.get(code_ref, "FFFFFF")
        row_fill = PatternFill("solid", fgColor=fill_color)

        for col_idx, val in enumerate(row[: len(HEADERS)], start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val if val != "" else None)
            cell.fill = row_fill

    # --- Largeurs de colonnes ---
    largeurs = {
        "graphique_code": 22,
        "thematique_code": 18,
        "type_graphique": 20,
        "serie_code": 16,
        "code_x": 12,
        "valeur": 10,
    }
    for i, h in enumerate(HEADERS, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = largeurs.get(h, 14)

    # --- Figer la première ligne ---
    ws.freeze_panes = "A2"

    output = Path(__file__).resolve().parent.parent.parent / "exemple_import.xlsx"
    wb.save(output)
    print(f"Fichier généré : {output}")


if __name__ == "__main__":
    generer()
