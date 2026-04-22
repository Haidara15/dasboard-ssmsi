from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


def build_thematique_update_template(thematique):
    wb = Workbook()
    ws = wb.active
    ws.title = "import"

    headers = [
        "graphique_code",
        "thematique_code",
        "type_graphique",
        "serie_code",
        "code_x",
        "valeur",
    ]

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF")

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    row_idx = 2

    graphiques = thematique.graphiques.all().order_by("ordre", "id")
    for graphique in graphiques:
        series = graphique.series.all().order_by("ordre", "id")

        for serie in series:
            observations = graphique.observations.filter(serie=serie).order_by("ordre_x", "id")

            if observations.exists():
                for obs in observations:
                    ws.append([
                        graphique.code,
                        thematique.code,
                        graphique.type_graphique,
                        serie.code,
                        obs.code_x,
                        obs.valeur,
                    ])
                    row_idx += 1
            else:
                ws.append([
                    graphique.code,
                    thematique.code,
                    graphique.type_graphique,
                    serie.code,
                    "",
                    "",
                ])
                row_idx += 1

    widths = {
        "A": 26,
        "B": 20,
        "C": 20,
        "D": 20,
        "E": 16,
        "F": 12,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    ws.freeze_panes = "A2"

    guide = wb.create_sheet("README")
    guide.append(["Consignes"]) 
    guide.append(["1. Ne pas modifier les colonnes de codes."])
    guide.append(["2. Mettre a jour uniquement la colonne valeur."])
    guide.append(["3. Vous pouvez ajouter de nouvelles lignes si necessaire."])
    guide.append(["4. Ne pas changer type_graphique."])

    return wb
