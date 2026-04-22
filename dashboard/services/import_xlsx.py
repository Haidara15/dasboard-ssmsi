"""
Service d'import xlsx -> base de donnees.

Version corrigee et commentee.

Principes :
- Lecture d'un fichier Excel (.xlsx)
- Creation / mise a jour des thematiques et graphiques
- Validation metier avant ecriture
- Remplacement complet des series / observations des graphiques presents dans le fichier
- Gestion des types a serie unique :
    camembert, anneau, valeur_cle
- Si plusieurs serie_code distincts pour ces types : erreur explicite
"""

import openpyxl
from collections import defaultdict

from dashboard.models import (
    Thematique,
    Graphique,
    SerieGraphique,
    Observation,
    TYPES_SERIE_UNIQUE,
)


# =========================================================
# Helpers generiques
# =========================================================

def _str(val):
    if val is None:
        return ""
    return str(val).strip()


def _bool(val, default=False):
    if val is None:
        return default
    if isinstance(val, bool):
        return val
    return str(val).strip().lower() in ("true", "1", "oui", "yes")


def _int(val, default=0):
    try:
        return int(val)
    except (TypeError, ValueError):
        return default


def _float(val):
    try:
        return float(val)
    except (TypeError, ValueError):
        return None


def _optional_str(r, key, current=""):
    val = r.get(key)
    if val is None:
        return current

    txt = str(val).strip()
    if txt == "":
        return current

    return txt


def _optional_bool(r, key, current=False):
    val = r.get(key)
    if val is None:
        return current

    if isinstance(val, str) and not val.strip():
        return current

    return _bool(val, default=current)


def _optional_int(r, key, current=0):
    val = r.get(key)
    if val is None:
        return current

    if isinstance(val, str) and not val.strip():
        return current

    return _int(val, default=current)


def _headers(sheet):
    return {
        str(cell.value).strip().lower(): idx
        for idx, cell in enumerate(next(sheet.iter_rows(min_row=1, max_row=1)))
        if cell.value is not None
    }


def _row_dict(headers, row):
    return {col: row[idx].value for col, idx in headers.items()}


# =========================================================
# Gestion thematique
# =========================================================

def _get_or_create_thematique(r, graphique_code, rapport):
    thematique_code = _str(r.get("thematique_code")) or _str(r.get("page_code"))

    if not thematique_code:
        rapport.append(f"{graphique_code} ignore : thematique_code manquant")
        return None

    thematique = Thematique.objects.filter(code=thematique_code).first()
    if thematique:
        return thematique

    titre = _str(r.get("thematique_titre")) or _str(r.get("page_titre"))

    if not titre:
        titre = thematique_code.replace("-", " ").title()

    thematique, _ = Thematique.objects.get_or_create(
        code=thematique_code,
        defaults={
            "titre": titre,
            "actif": True,
        }
    )

    rapport.append(f"Thematique creee : {thematique_code}")
    return thematique


# =========================================================
# Fonction principale
# =========================================================

def importer_xlsx(fichier):
    rapport = []

    # -----------------------------------------------------
    # Ouverture fichier
    # -----------------------------------------------------
    try:
        wb = openpyxl.load_workbook(fichier, read_only=True, data_only=True)
    except Exception as e:
        return [f"Impossible d'ouvrir le fichier : {e}"]

    sheet = wb.worksheets[0]
    rapport.append(f"Feuille lue : {sheet.title}")

    headers = _headers(sheet)

    colonnes_minimales = {
        "graphique_code",
        "serie_code",
        "code_x",
        "valeur",
    }

    manquantes = colonnes_minimales - set(headers.keys())

    if manquantes:
        return rapport + [
            f"Colonnes obligatoires manquantes : {', '.join(sorted(manquantes))}"
        ]

    # -----------------------------------------------------
    # Cache graphiques existants
    # -----------------------------------------------------
    graphiques_existants = {
        g.code: g
        for g in Graphique.objects.select_related("thematique").all()
    }

    # -----------------------------------------------------
    # Passe 1 : validation technique des lignes
    # -----------------------------------------------------
    rows_data = []
    lignes_invalides = 0

    for row_idx, row in enumerate(sheet.iter_rows(min_row=2), start=2):
        r = _row_dict(headers, row)

        if not any(v for v in r.values() if v is not None):
            continue

        graphique_code = _str(r.get("graphique_code"))
        serie_code = _str(r.get("serie_code"))
        code_x = _str(r.get("code_x"))
        valeur = r.get("valeur")
        thematique_code = _str(r.get("thematique_code"))
        type_graphique = _str(r.get("type_graphique"))
        graphique_existant = graphiques_existants.get(graphique_code)

        # Validation stricte selon le type de graphique
        # --- Camembert ---
        if type_graphique == "camembert":
            colonnes_oblig = ["graphique_code", "thematique_code", "type_graphique", "serie_code", "code_x", "valeur"]
            for col in colonnes_oblig:
                if not _str(r.get(col)):
                    rapport.append(f"Ligne {row_idx}: colonne obligatoire '{col}' manquante pour camembert. Import bloque.")
                    lignes_invalides += 1
                    break
            if not isinstance(valeur, (int, float)):
                try:
                    float(valeur)
                except Exception:
                    rapport.append(f"Ligne {row_idx}: valeur non numérique pour camembert. Import bloque.")
                    lignes_invalides += 1
                    continue

        # --- Histogramme ---
        elif type_graphique == "histogramme":
            colonnes_oblig = ["graphique_code", "thematique_code", "type_graphique", "serie_code", "code_x", "valeur"]
            for col in colonnes_oblig:
                if not _str(r.get(col)):
                    rapport.append(f"Ligne {row_idx}: colonne obligatoire '{col}' manquante pour histogramme. Import bloque.")
                    lignes_invalides += 1
                    break
            if not isinstance(valeur, (int, float)):
                try:
                    float(valeur)
                except Exception:
                    rapport.append(f"Ligne {row_idx}: valeur non numérique pour histogramme. Import bloque.")
                    lignes_invalides += 1
                    continue

        # --- Courbe ---
        elif type_graphique == "courbe":
            colonnes_oblig = ["graphique_code", "thematique_code", "type_graphique", "serie_code", "code_x", "valeur"]
            for col in colonnes_oblig:
                if not _str(r.get(col)):
                    rapport.append(f"Ligne {row_idx}: colonne obligatoire '{col}' manquante pour courbe. Import bloque.")
                    lignes_invalides += 1
                    break
            if not isinstance(valeur, (int, float)):
                try:
                    float(valeur)
                except Exception:
                    rapport.append(f"Ligne {row_idx}: valeur non numérique pour courbe. Import bloque.")
                    lignes_invalides += 1
                    continue

        # --- Défaut ou autre type ---
        else:
            # Validation générique minimale
            if not graphique_code or not serie_code or not code_x or valeur is None:
                lignes_invalides += 1
                continue
            try:
                float(valeur)
            except Exception:
                rapport.append(f"Ligne {row_idx}: valeur non numérique. Import bloque.")
                lignes_invalides += 1
                continue

        # Si tout est ok, on ajoute la ligne
        rows_data.append(r)


    if lignes_invalides > 0:
        rapport.append(f"Import interrompu : {lignes_invalides} ligne(s) invalide(s). Corrigez le fichier et recommencez.")
        return rapport

    if not rows_data:
        return rapport + [f"Aucune ligne valide ({lignes_invalides} rejetees)"]

    # -----------------------------------------------------
    # Passe 2 : validation métier (AUCUNE écriture en base avant cette étape)
    # -----------------------------------------------------
    groupes = defaultdict(list)

    for r in rows_data:
        graphique_code = _str(r.get("graphique_code"))
        groupes[graphique_code].append(r)

    for graphique_code, lignes in groupes.items():
        # On ne touche pas à la base ici, on ne fait que valider
        graphique_existant = graphiques_existants.get(graphique_code)

        type_graphique = _str(lignes[0].get("type_graphique"))

        if not type_graphique and graphique_existant:
            type_graphique = graphique_existant.type_graphique

        type_graphique = type_graphique.lower()

        if type_graphique in TYPES_SERIE_UNIQUE:
            series_uniques = {
                _str(l.get("serie_code"))
                for l in lignes
                if _str(l.get("serie_code"))
            }

            if len(series_uniques) > 1:
                return rapport + [
                    (
                        f"Erreur : le graphique '{graphique_code}' "
                        f"de type '{type_graphique}' "
                        f"n'accepte qu'une seule serie. "
                        f"Series detectees : {', '.join(sorted(series_uniques))}"
                    )
                ]

    # -----------------------------------------------------
    # PASSE 3+ : TOUTE LA CREATION EN BASE DANS UNE TRANSACTION
    # -----------------------------------------------------

    # -----------------------------------------------------
    # PASSE 3+ : TOUTE LA CREATION EN BASE DANS UNE TRANSACTION
    # -----------------------------------------------------
    from django.db import transaction, connection

    try:
        with transaction.atomic():
            graphiques_traites = {}
            thematiques_cache = {}
            graphiques_crees = 0
            graphiques_maj = 0

            # Création/mise à jour des thématiques et graphiques
            for r in rows_data:
                graphique_code = _str(r.get("graphique_code"))

                if graphique_code in graphiques_traites:
                    continue

                graphique_existant = graphiques_existants.get(graphique_code)

                graphique_titre = _str(r.get("graphique_titre"))

                if not graphique_titre and graphique_existant:
                    graphique_titre = graphique_existant.titre

                if not graphique_titre:
                    graphique_titre = graphique_code.replace("-", " ").title()

                type_graphique = _str(r.get("type_graphique"))
                if not type_graphique and graphique_existant:
                    type_graphique = graphique_existant.type_graphique

                thematique_code = _str(r.get("thematique_code"))
                if not thematique_code and graphique_existant:
                    thematique_code = graphique_existant.thematique.code

                if thematique_code not in thematiques_cache:
                    thematiques_cache[thematique_code] = _get_or_create_thematique(
                        r, graphique_code, rapport
                    )

                thematique = thematiques_cache[thematique_code]

                graphique, created = Graphique.objects.update_or_create(
                    code=graphique_code,
                    defaults={
                        "titre": graphique_titre,
                        "type_graphique": type_graphique,
                        "thematique": thematique,
                        "sous_titre": _optional_str(
                            r, "graphique_sous_titre",
                            current=graphique_existant.sous_titre if graphique_existant else ""
                        ),
                        "titre_axe_x": _optional_str(
                            r, "titre_axe_x",
                            current=graphique_existant.titre_axe_x if graphique_existant else ""
                        ),
                        "titre_axe_y": _optional_str(
                            r, "titre_axe_y",
                            current=graphique_existant.titre_axe_y if graphique_existant else ""
                        ),
                        "legende_visible": _optional_bool(
                            r, "legende_visible",
                            current=graphique_existant.legende_visible if graphique_existant else True
                        ),
                        "source": _optional_str(
                            r, "source",
                            current=graphique_existant.source if graphique_existant else ""
                        ),
                        "champ": _optional_str(
                            r, "champ",
                            current=graphique_existant.champ if graphique_existant else ""
                        ),
                        "note": _optional_str(
                            r, "note",
                            current=graphique_existant.note if graphique_existant else ""
                        ),
                        "texte_valeur_cle": _optional_str(
                            r, "texte_valeur_cle",
                            current=graphique_existant.texte_valeur_cle if graphique_existant else ""
                        ),
                        "ordre": _optional_int(
                            r, "ordre_graphique",
                            current=graphique_existant.ordre if graphique_existant else 0
                        ),
                        "actif": _optional_bool(
                            r, "actif",
                            current=graphique_existant.actif if graphique_existant else True
                        ),
                    }
                )

                graphiques_traites[graphique_code] = graphique

                if created:
                    graphiques_crees += 1
                else:
                    graphiques_maj += 1

            rapport.append(
                f"Graphiques : {graphiques_crees} crees, {graphiques_maj} mis a jour"
            )

            # Reset des series / observations
            ids_graph = [g.pk for g in graphiques_traites.values() if g]
            SerieGraphique.objects.filter(graphique_id__in=ids_graph).delete()

            # Resync SQLite sequence après delete
            with connection.cursor() as c:
                c.execute(
                    "UPDATE sqlite_sequence SET seq = (SELECT COALESCE(MAX(id), 0) FROM dashboard_seriegraphique) WHERE name = 'dashboard_seriegraphique'"
                )

            # Création des séries et observations
            series_cache = {}
            ordre_x_counter = {}
            obs_crees = 0

            for r in rows_data:
                graphique_code = _str(r.get("graphique_code"))
                graphique = graphiques_traites.get(graphique_code)

                if not graphique:
                    continue

                serie_code = _str(r.get("serie_code"))
                code_x = _str(r.get("code_x"))
                valeur = _float(r.get("valeur"))

                cache_key = (graphique.pk, serie_code)

                if cache_key not in series_cache:
                    serie_libelle = (
                        _str(r.get("serie_libelle"))
                        or serie_code.replace("-", " ").title()
                    )

                    serie = SerieGraphique.objects.create(
                        graphique=graphique,
                        code=serie_code,
                        libelle=serie_libelle,
                        couleur=_str(r.get("serie_couleur")),
                        ordre=_int(r.get("serie_ordre"), default=0),
                        visible=_bool(r.get("serie_visible"), default=True),
                    )

                    series_cache[cache_key] = serie

                serie = series_cache[cache_key]

                if r.get("ordre_x") is not None and _str(r.get("ordre_x")) != "":
                    ordre_x = _int(r.get("ordre_x"))
                else:
                    ordre_x_counter[graphique_code] = (
                        ordre_x_counter.get(graphique_code, 0) + 1
                    )
                    ordre_x = ordre_x_counter[graphique_code]

                Observation.objects.create(
                    graphique=graphique,
                    serie=serie,
                    code_x=code_x,
                    libelle_x=_optional_str(r, "libelle_x", current=code_x),
                    ordre_x=ordre_x,
                    valeur=valeur,
                    valeur_formatee=_optional_str(r, "valeur_formatee"),
                    infobulle=_optional_str(r, "infobulle"),
                )

                obs_crees += 1

            rapport.append(f"Observations : {obs_crees} creees")

    except Exception as e:
        return rapport + [f"Erreur : {e}"]

    return rapport